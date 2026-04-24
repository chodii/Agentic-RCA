# -*- coding: utf-8 -*-
"""
Given:
1) the XLSX workbook
2) the JSON produced by xlsx_extract_dates_positions.py
3) the matches JSON produced by plot_json_dates_histogram.py

extract the FULL ROWS from the workbook for rows whose date cells belong to
matched DATE-ONLY overlaps.

This version trims trailing empty columns so you do not get long tails of null
headers / values.

Usage:
    python ./xlsx_extract_matched_rows.py ^
        "C:/path/to/file.xlsx" ^
        "C:/path/to/xlsx_date_positions.json" ^
        "C:/path/to/json_date_matches.json"

Optional:
    python ./xlsx_extract_matched_rows.py ^
        "C:/path/to/file.xlsx" ^
        "C:/path/to/xlsx_date_positions.json" ^
        "C:/path/to/json_date_matches.json" ^
        --out matched_rows.json ^
        --include-empty
"""

import argparse
import json
from collections import defaultdict
from datetime import date, datetime
from typing import Any, Dict, List, Set

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def parse_date_only(s: str) -> date:
    s = s.strip()
    return date.fromisoformat(s)


def parse_timestamp_to_date(s: str) -> date:
    s = s.strip()
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    try:
        if "T" in s or " " in s:
            return datetime.fromisoformat(s).date()
        return date.fromisoformat(s)
    except ValueError as e:
        raise ValueError(f"Unsupported timestamp/date format: {s}") from e


def load_matched_dates(matches_json_path: str) -> Set[date]:
    with open(matches_json_path, "r", encoding="utf-8") as f:
        obj = json.load(f)

    overlaps = obj.get("date_only_overlaps", [])
    if not isinstance(overlaps, list):
        raise ValueError('"date_only_overlaps" must be a list in matches JSON.')

    matched_dates = set()
    for s in overlaps:
        if isinstance(s, str):
            matched_dates.add(parse_date_only(s))
    return matched_dates


def load_target_rows_from_positions(positions_json_path: str, matched_dates: Set[date]):
    with open(positions_json_path, "r", encoding="utf-8") as f:
        obj = json.load(f)

    if not isinstance(obj, dict):
        raise ValueError("Positions JSON must be a dict mapping timestamps to position lists.")

    targets_by_sheet_name: Dict[str, Dict[int, Dict[str, Any]]] = defaultdict(dict)

    for timestamp_key, positions in obj.items():
        if not isinstance(timestamp_key, str):
            continue

        ts_date = parse_timestamp_to_date(timestamp_key)
        if ts_date not in matched_dates:
            continue

        if not isinstance(positions, list):
            continue

        for pos in positions:
            if not isinstance(pos, dict):
                continue

            sheet_name = pos.get("sheet_name")
            sheet_id = pos.get("sheet_id")
            row = pos.get("row")
            col = pos.get("column")

            if not isinstance(sheet_name, str):
                continue
            if not isinstance(row, int):
                continue

            entry = targets_by_sheet_name[sheet_name].setdefault(
                row,
                {
                    "sheet_id": sheet_id,
                    "matched_dates": set(),
                    "matched_date_cells": set(),
                },
            )
            entry["matched_dates"].add(ts_date.isoformat())
            if isinstance(col, str):
                entry["matched_date_cells"].add(f"{col}{row}")

    return targets_by_sheet_name


def json_safe_value(value):
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def is_empty_value(value: Any) -> bool:
    return value is None or value == ""


def trim_trailing_empty(values: List[Any]) -> List[Any]:
    end = len(values)
    while end > 0 and is_empty_value(values[end - 1]):
        end -= 1
    return values[:end]


def extract_rows(
    xlsx_path: str,
    targets_by_sheet_name,
    include_empty: bool = False,
):
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    out_rows: List[Dict[str, Any]] = []

    wanted_sheet_names = set(targets_by_sheet_name.keys())

    for sheet_id, ws in enumerate(wb.worksheets, start=1):
        if ws.title not in wanted_sheet_names:
            continue

        sheet_targets = targets_by_sheet_name[ws.title]
        wanted_rows = set(sheet_targets.keys())
        if not wanted_rows:
            continue

        headers = None
        remaining_rows = set(wanted_rows)

        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            raw_values = [json_safe_value(cell.value) for cell in row]

            if row_idx == 1:
                headers = trim_trailing_empty(raw_values)

            if row_idx not in wanted_rows:
                continue

            meta = sheet_targets[row_idx]

            row_values = trim_trailing_empty(raw_values)

            effective_len = max(
                len(row_values),
                len(headers) if headers is not None else 0,
            )

            row_values = row_values[:effective_len]
            if headers is not None:
                headers_for_row = headers[:effective_len]
                if len(headers_for_row) < effective_len:
                    headers_for_row = headers_for_row + [None] * (effective_len - len(headers_for_row))
            else:
                headers_for_row = []

            cells = []
            for col_idx in range(1, effective_len + 1):
                value = row_values[col_idx - 1] if col_idx - 1 < len(row_values) else None

                if is_empty_value(value) and not include_empty:
                    continue

                header = headers_for_row[col_idx - 1] if col_idx - 1 < len(headers_for_row) else None

                cells.append(
                    {
                        "column": get_column_letter(col_idx),
                        "header": header,
                        "value": value,
                    }
                )

            out_rows.append(
                {
                    "sheet_id": meta["sheet_id"] if isinstance(meta.get("sheet_id"), int) else sheet_id,
                    "sheet_name": ws.title,
                    "row": row_idx,
                    "matched_dates": sorted(meta["matched_dates"]),
                    "matched_date_cells": sorted(meta["matched_date_cells"]),
                    "headers": headers_for_row,
                    "row_values": row_values,
                    "cells": cells,
                }
            )

            if row_idx in remaining_rows:
                remaining_rows.remove(row_idx)
            if not remaining_rows:
                break

    wb.close()
    return out_rows


def main():
    parser = argparse.ArgumentParser(
        description="Extract full XLSX rows based on matched dates and date-cell positions."
    )
    parser.add_argument("xlsx_path", help="Path to the workbook (.xlsx)")
    parser.add_argument("positions_json", help="Path to xlsx_date_positions.json")
    parser.add_argument("matches_json", help="Path to json_date_matches.json")
    parser.add_argument(
        "--out",
        default="matched_rows.json",
        help="Output JSON path (default: matched_rows.json)",
    )
    parser.add_argument(
        "--include-empty",
        action="store_true",
        help="Include empty cells in the structured 'cells' output.",
    )
    args = parser.parse_args()

    matched_dates = load_matched_dates(args.matches_json)
    if not matched_dates:
        print("No date-only overlaps found in matches JSON.")
        with open(args.out, "w", encoding="utf-8") as f:
            json.dump([], f, ensure_ascii=False, indent=2)
        print(f"Saved empty result to: {args.out}")
        return

    targets_by_sheet_name = load_target_rows_from_positions(args.positions_json, matched_dates)
    results = extract_rows(
        args.xlsx_path,
        targets_by_sheet_name,
        include_empty=args.include_empty,
    )

    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)

    print(f"Matched dates: {len(matched_dates)}")
    print(f"Extracted rows: {len(results)}")
    print(f"Saved rows to: {args.out}")


if __name__ == "__main__":
    main()
