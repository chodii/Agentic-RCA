# -*- coding: utf-8 -*-
"""
Extract all datetime-like cells from an XLSX workbook and save only:
- normalized datetime
- sheet id / sheet name
- row
- column

Usage:
    python .\\xlsx_extract_dates_positions.py "C:\\path\\to\\file.xlsx"
    python .\\xlsx_extract_dates_positions.py "C:\\path\\to\\file.xlsx" --out "dates_positions.json"
"""

import argparse
import json
import re
from datetime import date, datetime, time
from typing import Optional

from openpyxl import load_workbook
from openpyxl.styles.numbers import is_date_format
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel


DATEISH_RE = re.compile(
    r"""^
    \s*
    (?:
        \d{1,2}/\d{1,2}/\d{2,4}      # 8/18/2025 or 8/18/25
        (?:\s+\d{1,2}:\d{2}(?::\d{2})?\s*(?:AM|PM)?)?
      |
        \d{1,2}-[A-Za-z]{3}-\d{2,4}  # 29-Mar-21
        (?:\s+\d{1,2}:\d{2}(?::\d{2})?\s*(?:AM|PM)?)?
    )
    \s*$
    """,
    re.IGNORECASE | re.VERBOSE,
)


TEXT_DT_FORMATS = [
    "%m/%d/%Y %I:%M:%S %p",
    "%m/%d/%Y %I:%M %p",
    "%m/%d/%Y %H:%M:%S",
    "%m/%d/%Y %H:%M",
    "%m/%d/%Y",
    "%m/%d/%y %I:%M:%S %p",
    "%m/%d/%y %I:%M %p",
    "%m/%d/%y %H:%M:%S",
    "%m/%d/%y %H:%M",
    "%m/%d/%y",
    "%d-%b-%Y %H:%M:%S",
    "%d-%b-%Y %H:%M",
    "%d-%b-%Y",
    "%d-%b-%y %H:%M:%S",
    "%d-%b-%y %H:%M",
    "%d-%b-%y",
]


def looks_like_date_string(s: str) -> bool:
    s = s.strip()
    if not s or len(s) > 40:
        return False
    return bool(DATEISH_RE.match(s))


def parse_datetime_text(s: str) -> Optional[datetime]:
    s = s.strip()
    for fmt in TEXT_DT_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def normalize_datetime_key(dt: datetime) -> str:
    dt = dt.replace(microsecond=0, tzinfo=None)
    return dt.isoformat()


def try_extract_datetime(cell, workbook_epoch) -> Optional[datetime]:
    value = cell.value

    if value is None:
        return None

    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return datetime.combine(value, time.min)

    if isinstance(value, (int, float)):
        try:
            if is_date_format(getattr(cell, "number_format", "")):
                converted = from_excel(value, epoch=workbook_epoch)
                if isinstance(converted, datetime):
                    return converted
                if isinstance(converted, date):
                    return datetime.combine(converted, time.min)
        except Exception:
            return None

    if isinstance(value, str):
        if looks_like_date_string(value):
            return parse_datetime_text(value)

    return None


def extract_date_positions(xlsx_path: str):
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    out = {}

    for sheet_id, ws in enumerate(wb.worksheets, start=1):
        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            for col_idx, cell in enumerate(row, start=1):
                try:
                    dt = try_extract_datetime(cell, wb.epoch)
                except Exception:
                    dt = None

                if dt is None:
                    continue

                key = normalize_datetime_key(dt)
                out.setdefault(key, []).append({
                    "sheet_id": sheet_id,
                    "sheet_name": ws.title,
                    "row": row_idx,
                    "column": get_column_letter(col_idx),
                })

    wb.close()
    return out


def main():
    parser = argparse.ArgumentParser(
        description="Extract datetime-like cells from XLSX and save their positions to JSON."
    )
    parser.add_argument("xlsx_path", help="Path to the .xlsx file")
    parser.add_argument(
        "--out",
        default="xlsx_date_positions.json",
        help="Output JSON path (default: xlsx_date_positions.json)",
    )
    args = parser.parse_args()

    results = extract_date_positions(args.xlsx_path)

    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, separators=(",", ":"))

    num_dates = len(results)
    num_positions = sum(len(v) for v in results.values())
    print(f"Saved {num_positions} positions for {num_dates} unique datetimes to: {args.out}")


if __name__ == "__main__":
    main()

